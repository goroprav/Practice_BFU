import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel, QLineEdit, QTextEdit, QFileDialog, 
                             QMessageBox, QVBoxLayout, QHBoxLayout, QWidget, QSpacerItem, QSizePolicy)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from data_loader import DataLoader
from data_processor import DataProcessor

class DataApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Обработчик")
        self.setGeometry(100, 100, 900, 600)
        
        self.setStyleSheet("background-color: #EAEAFC;")  # Задний фон светло-голубого цвета

        self.data = None
        self.processed_data_x = None
        self.processed_data_y = None

        main_layout = QHBoxLayout()
        input_layout = QVBoxLayout()
        output_layout = QVBoxLayout()

        self.upload_button = QPushButton("Загрузить файл", self)
        self.upload_button.clicked.connect(self.upload_file)
        self.upload_button.setStyleSheet(self.get_button_style())
        
        input_layout.addWidget(self.upload_button)

        self.button_x1_y1 = QPushButton("Пересчет Peak information", self)
        self.button_x1_y1.setCheckable(True)
        self.button_x1_y1.clicked.connect(self.select_peak_information)
        self.button_x1_y1.setStyleSheet(self.get_toggle_button_style(False))
        
        input_layout.addWidget(self.button_x1_y1)

        self.button_x2_y2 = QPushButton("Пересчет Plot Curves", self)
        self.button_x2_y2.setCheckable(True)
        self.button_x2_y2.clicked.connect(self.select_plot_curves)
        self.button_x2_y2.setStyleSheet(self.get_toggle_button_style(False))
        
        input_layout.addWidget(self.button_x2_y2)

        self.coefficient_1_label = QLabel("Коэффициент a:", self)
        self.coefficient_1_label.setStyleSheet("color: black;")
        input_layout.addWidget(self.coefficient_1_label)
        self.coefficient_1_entry = QLineEdit(self)
        self.coefficient_1_entry.setText("0.0000083526")  # Установить начальное значение
        self.coefficient_1_entry.setStyleSheet(self.get_line_edit_style())
        input_layout.addWidget(self.coefficient_1_entry)

        self.coefficient_2_label = QLabel("Коэффициент b:", self)
        self.coefficient_2_label.setStyleSheet("color: black;")
        input_layout.addWidget(self.coefficient_2_label)
        self.coefficient_2_entry = QLineEdit(self)
        self.coefficient_2_entry.setText("0.98")  # Установить начальное значение
        self.coefficient_2_entry.setStyleSheet(self.get_line_edit_style())
        input_layout.addWidget(self.coefficient_2_entry)

        self.lambda_0_label = QLabel("Длина волны:", self)
        self.lambda_0_label.setStyleSheet("color: black;")
        self.lambda_0_entry = QLineEdit(self)
        self.lambda_0_entry.setText("532")  # Установить начальное значение
        self.lambda_0_entry.setStyleSheet(self.get_line_edit_style())
        self.temp_T_label = QLabel("Температура:", self)
        self.temp_T_label.setStyleSheet("color: black;")
        self.temp_T_entry = QLineEdit(self)
        self.temp_T_entry.setText("273")  # Установить начальное значение
        self.temp_T_entry.setStyleSheet(self.get_line_edit_style())

        input_layout.addWidget(self.lambda_0_label)
        input_layout.addWidget(self.lambda_0_entry)
        input_layout.addWidget(self.temp_T_label)
        input_layout.addWidget(self.temp_T_entry)

        self.lambda_0_label.hide()
        self.lambda_0_entry.hide()
        self.temp_T_label.hide()
        self.temp_T_entry.hide()

        spacer = QSpacerItem(12, 20, QSizePolicy.Minimum, QSizePolicy.Expanding)
        input_layout.addItem(spacer)

        self.process_button = QPushButton("Обработать данные", self)
        self.process_button.clicked.connect(self.process_data)
        self.process_button.setStyleSheet(self.get_button_style())
       
        output_layout.addWidget(self.process_button)

        self.output_text = QTextEdit(self)
        self.output_text.setStyleSheet(self.get_output_text_style())
        output_layout.addWidget(self.output_text)

        self.save_button = QPushButton("Сохранить в Excel", self)
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setStyleSheet(self.get_button_style())
      
        output_layout.addWidget(self.save_button)

        main_layout.addLayout(input_layout, 1)
        main_layout.addLayout(output_layout, 3)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

    def upload_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Загрузить файл", "", "All Files (*)")
        if file_path:
            loader = DataLoader(file_path)
            self.data = loader.load_data()
            QMessageBox.information(self, "Загрузка файла", "Файл успешно загружен")
            self.temp_T_label.setStyleSheet("color: black;")

    def select_peak_information(self):
        if not self.data:
            QMessageBox.critical(self, "Ошибка", "Сначала загрузите файл")
            self.button_x1_y1.setChecked(False)
            return
        self.button_x1_y1.setChecked(True)
        self.button_x1_y1.setStyleSheet(self.get_toggle_button_style(False))
        self.button_x2_y2.setChecked(False)
        self.button_x2_y2.setStyleSheet(self.get_toggle_button_style(False))
        self.select_data_set("Пересчет Peak information")

    def select_plot_curves(self):
        if not self.data:
            QMessageBox.critical(self, "Ошибка", "Сначала загрузите файл")
            self.button_x2_y2.setChecked(False)
            return
        self.button_x2_y2.setChecked(True)
        self.button_x2_y2.setStyleSheet(self.get_toggle_button_style(False))
        self.button_x1_y1.setChecked(False)
        self.button_x1_y1.setStyleSheet(self.get_toggle_button_style(False))
        self.select_data_set("Пересчет Plot Curves")

    def select_data_set(self, selected):
        self.output_text.clear()
        if selected == "Пересчет Peak information":
            data_x = self.data[0]
            for x in data_x:
                self.output_text.append(f"{x}")

            self.lambda_0_label.hide()
            self.lambda_0_entry.hide()
            self.temp_T_label.hide()
            self.temp_T_entry.hide()
        else:
            data_x = self.data[2]
            data_y = self.data[3]
            for x, y in zip(data_x, data_y):
                self.output_text.append(f"{x}\t{y}")

            self.lambda_0_label.show()
            self.lambda_0_entry.show()
            self.temp_T_label.show()
            self.temp_T_entry.show()

    def process_data(self):
        if not self.data:
            QMessageBox.critical(self, "Ошибка", "Сначала загрузите файл")
            return

        selected_data = "Пересчет Peak information" if self.button_x1_y1.isChecked() else "Пересчет Plot Curves"
        try:
            coefficient_1 = float(self.coefficient_1_entry.text())
            coefficient_2 = float(self.coefficient_2_entry.text())
            if selected_data == "Пересчет Peak information":
                data_x = self.data[0]
                processor = DataProcessor(data_x, coefficient_1=coefficient_1, coefficient_2=coefficient_2)
                self.processed_data_x = processor.process_x1()
                self.processed_data_y = None
            else:
                data_x = self.data[2]
                data_y = self.data[3]
                lambda_0 = float(self.lambda_0_entry.text()) * 10 ** (-9)
                temp_T = float(self.temp_T_entry.text())
                processor = DataProcessor(data_x, data_y, coefficient_1, coefficient_2, lambda_0, temp_T)
                self.processed_data_x, self.processed_data_y = processor.process_x2_y2()
        except ValueError:
            QMessageBox.critical(self, "Ошибка", "Введите корректные значения для коэффициентов")
            return

        self.output_text.clear()
        if selected_data == "Пересчет Peak information":
            for x, new_x in zip(self.data[0], self.processed_data_x):
                self.output_text.append(f"{x}\t\t{new_x}")
        else:
            for x, y, new_x, new_y in zip(self.data[2], self.data[3], self.processed_data_x, self.processed_data_y):
                self.output_text.append(f"{x}\t{y}\t{new_x}\t{new_y}")

    def save_to_excel(self):
        if self.processed_data_x is None:
            QMessageBox.critical(self, "Ошибка", "Нет данных для сохранения")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Excel Files (*.xlsx)")
        if file_path:
            if self.button_x1_y1.isChecked():
                df = pd.DataFrame({
                    "X": self.data[0],
                    "Y": self.data[1],
                    
                    "Корректированная X": self.processed_data_x
                })
                coefficients = pd.DataFrame({
                    "Coefficient a": [float(self.coefficient_1_entry.text())],
                    "Coefficient b": [float(self.coefficient_2_entry.text())]
                })
            else:
                df = pd.DataFrame({
                    "X": self.data[2],
                    "Y": self.data[3],
                    "Корректированная X": self.processed_data_x,
                    "Корректированная Y": self.processed_data_y
                })
                coefficients = pd.DataFrame({
                    "Coefficient 1": [float(self.coefficient_1_entry.text())],
                    "Coefficient 2": [float(self.coefficient_2_entry.text())],
                    "Lambda 0": [float(self.lambda_0_entry.text())],
                    "Temp T": [float(self.temp_T_entry.text())]
                })

            with pd.ExcelWriter(file_path) as writer:
                df.to_excel(writer, sheet_name='Processed Data', index=False)
                coefficients.to_excel(writer, sheet_name='Coefficients', index=False)
                if self.button_x2_y2.isChecked():  # Добавить массив Y1
                    pd.DataFrame({"Y1": self.data[3]}).to_excel(writer, sheet_name='Original Y1', index=False)

            self.adjust_column_widths(file_path, 'Processed Data')
            self.adjust_column_widths(file_path, 'Coefficients')
            if self.button_x2_y2.isChecked():
                self.adjust_column_widths(file_path, 'Original Y1')

            QMessageBox.information(self, "Сохранение файла", "Файл успешно сохранен")

    def adjust_column_widths(self, file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 2
            sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(file_path)


    def get_button_style(self):
        return """
        QPushButton {
            background-color: #21145F;
            color: white;
            border-radius: 10px;
            padding: 10px;
            min-width: 160px;
        }
        QPushButton:hover {
            background-color: #1A1047;
        }
        """

    def get_toggle_button_style(self, active):
        if active:
            return """
            QPushButton {
                background-color: #21145F;
                color: white;
                border-radius: 10px;
                min-width: 160px;
            }
            QPushButton:hover {
                background-color: #21145F;
                color: white;
                border-radius: 10px;
                min-width: 160px;
            }
            """
        else:
            return """
            QPushButton {
                background-color: white;
                color: black;
                border-radius: 10px;
                padding: 10px;
                min-width: 160px;
            }
            QPushButton:hover {
                background-color: #21145F;
                color: white;
                border-radius: 10px;
                padding: 10px;
            }
            """

    def get_line_edit_style(self):
        return """
        QLineEdit {
            background-color: white;
            border-radius: 10px;
            color: black;
            padding: 6px;
        }
        QLineEdit:focus {
            background-color: white;
            border-radius: 10px;
            color: black;
            padding: 5px;
        }
        """

    def get_output_text_style(self):
        return """
        QTextEdit {
            background-color: white;
            border-radius: 10px;
            color: black;
            padding: 10px;
            font-size: 12px;
            tab-width: 2;
        }
        QTextEdit QScrollBar:vertical {
            background-color: white;
            width: 14px;
            color: white;
            border-radius: 7px;
        }
        QTextEdit QScrollBar::handle:vertical {
            background-color: #21145F;
            min-height: 20px;
            border-radius: 7px;
        }
        QTextEdit QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            background-color: white;
        }
        QTextEdit QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background-color: white;
        }
        QTextEdit QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background-color: white;
        }
        """

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = DataApp()
    window.show()
    sys.exit(app.exec_())
